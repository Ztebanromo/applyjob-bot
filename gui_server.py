import atexit
import collections
import json
import logging as _logging
import os
import subprocess
import threading
import time
import sys
import secrets
import re
import tempfile
import signal
from functools import wraps
from flask import Flask, render_template, request, jsonify, Response as FlaskResponse, make_response

_log = _logging.getLogger("applyjob.server")


# ── Rate limiter ──────────────────────────────────────────────────────────────
class _RateLimiter:
    """
    Sliding-window rate limiter. Thread-safe.
    Raises RuntimeError if more than max_calls happen within window_s seconds.
    """
    def __init__(self, max_calls: int, window_s: float):
        self._max   = max_calls
        self._win   = window_s
        self._calls: collections.deque = collections.deque()
        self._lock  = threading.Lock()

    def check(self) -> None:
        now = time.time()
        with self._lock:
            while self._calls and now - self._calls[0] > self._win:
                self._calls.popleft()
            if len(self._calls) >= self._max:
                raise RuntimeError(
                    f"Rate limit: max {self._max} requests per {int(self._win)}s exceeded"
                )
            self._calls.append(now)


_config_rate_limiter = _RateLimiter(max_calls=10, window_s=60)
_env_write_lock = threading.Lock()  # protege escrituras concurrentes a .env
from flask_socketio import SocketIO, emit
from dotenv import load_dotenv

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', secrets.token_hex(24))

# ── HTTP Basic Auth (opcional) ────────────────────────────────────────────────
# Configura DASHBOARD_PASSWORD en .env para proteger el dashboard.
# Si está vacío, no pide autenticación (comportamiento por defecto).
_DASHBOARD_PW = os.getenv("DASHBOARD_PASSWORD", "")
if not _DASHBOARD_PW:
    print(
        "\n[AVISO DE SEGURIDAD] DASHBOARD_PASSWORD no está configurado.\n"
        "  Cualquier proceso local puede acceder al dashboard y disparar el bot.\n"
        "  Agrega DASHBOARD_PASSWORD=<clave> en .env para protegerlo.\n"
    )


# Nota: la autenticación se aplica globalmente via _global_auth() (before_request)
# No se usa ningún decorator individual — toda protección es centralizada.


# Inicializar SocketIO con hilos (modo más compatible en Windows sin eventlet/gevent)
# SECURITY: CORS restringido a localhost — nunca permitir orígenes externos
socketio = SocketIO(
    app,
    cors_allowed_origins=["http://127.0.0.1:5000", "http://localhost:5000"],
    async_mode='threading',
)

# Asegurar directorios
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


@app.before_request
def _global_auth():
    """
    Aplica HTTP Basic Auth a TODAS las rutas si DASHBOARD_PASSWORD está configurado.
    Las peticiones de SocketIO pasan porque van por WebSocket, no HTTP normal.
    """
    if not _DASHBOARD_PW:
        return
    auth = request.authorization
    if auth and auth.password == _DASHBOARD_PW:
        return
    return FlaskResponse(
        "Acceso restringido — configura DASHBOARD_PASSWORD en .env.",
        401,
        {"WWW-Authenticate": 'Basic realm="ApplyJob Dashboard"'},
    )

# ── Archivo señal de parada (compartido con engine.py) ────────────────────────
# gui_server escribe este archivo; engine.py lo detecta y cierra el browser limpio
_BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR        = os.path.join(_BASE_DIR, "data")
STOP_SIGNAL_PATH = os.path.join(_DATA_DIR, "STOP_SIGNAL")
os.makedirs(_DATA_DIR, exist_ok=True)

def _write_stop_signal():
    """Escribe el archivo señal para que engine.py cierre el browser limpiamente."""
    try:
        with open(STOP_SIGNAL_PATH, "w") as f:
            f.write(str(os.getpid()))
    except Exception:
        pass

def _clear_stop_signal():
    """Elimina el archivo señal tras confirmar que el proceso murió."""
    try:
        if os.path.exists(STOP_SIGNAL_PATH):
            os.remove(STOP_SIGNAL_PATH)
    except Exception:
        pass

# Limpiar señal residual al iniciar (por si el servidor se reinició abruptamente)
_clear_stop_signal()

# NOTA: scan_queue.json NO se limpia al arrancar.
# Las ofertas con preguntas desconocidas se guardan entre sesiones para que
# --apply-queue las reintente una vez el usuario haya respondido las preguntas.
# El bot aplica poda automática por antigüedad (5 días) en run_apply_queue.

# Max seconds a bot subprocess may run before the watchdog kills it
# Scan / apply-queue: 600s es suficiente (pocas ofertas, < 10 min)
# Master (--persistent / --multi-keyword): necesita horas → 7200s = 2 h
PORTAL_TIMEOUT_S       = int(os.getenv("PORTAL_TIMEOUT_S",        "600"))   # scan / cola
MASTER_TIMEOUT_S       = int(os.getenv("MASTER_TIMEOUT_S",       "7200"))   # run maestro (2h)


def _start_watchdog(process: subprocess.Popen, timeout_s: int, label: str) -> threading.Thread:
    """
    Lanza un thread daemon que mata `process` si no termina en timeout_s segundos.
    """
    def _watch():
        deadline = time.time() + timeout_s
        while time.time() < deadline:
            if process.poll() is not None:
                return          # proceso ya terminó — watchdog no necesario
            time.sleep(5)
        # Timeout agotado
        if process.poll() is None:
            _log.warning("[WATCHDOG] %s excedió %ds — matando PID %d",
                         label, timeout_s, process.pid)
            try:
                state.add_log(f"\n[WATCHDOG] ⏱ Timeout {timeout_s}s en {label} — forzando cierre.\n")
            except Exception:
                pass
            _write_stop_signal()
            time.sleep(3)
            if process.poll() is None:
                try:
                    process.kill()
                except Exception:
                    pass

    t = threading.Thread(target=_watch, daemon=True, name=f"watchdog-{label}")
    t.start()
    return t


def _kill_chromium_children(pid: int) -> None:
    """Mata todos los procesos Chromium hijos de pid usando psutil (si está disponible)."""
    try:
        import psutil
        parent = psutil.Process(pid)
        for child in parent.children(recursive=True):
            try:
                if "chrom" in child.name().lower():
                    child.kill()
                    _log.debug("[CLEANUP] Chromium hijo PID %d matado", child.pid)
            except Exception:
                pass
    except ImportError:
        pass   # psutil no instalado — omitir silenciosamente
    except Exception as e:
        _log.debug("[CLEANUP] Error limpiando hijos: %s", e)

# Estado global thread-safe
class BotState:
    def __init__(self):
        self.scan_process  = None   # subproceso de scan (independiente)
        self.apply_process = None   # subproceso de postulación (independiente)
        self.logs = []
        self.lock = threading.RLock()
        self.stop_requested = False
        self.scan_active  = False
        self.apply_active = False
        self.stats = {"applied": 0, "external": 0, "filtered": 0, "errors": 0, "no_nav": 0, "total": 0}
        self.current_portal = None
        self.intervention = None
        self.scan_run_id  = 0
        self.apply_run_id = 0

    @property
    def is_active(self):
        return self.scan_active or self.apply_active

    @property
    def process(self):
        """Compatibilidad con código legado."""
        return self.apply_process or self.scan_process

    def add_log(self, message):
        print(f"[BOT] {message.strip()}", flush=True)
        with self.lock:
            self.logs.append(message)
            if len(self.logs) > 2000:
                self.logs.pop(0)
        
        # Notificar a los clientes vía SocketIO
        socketio.emit('new_log', {"message": message}, namespace='/bot')
        
        # Detectar portal activo para actualizar la UI
        if "[PORTAL_ACTIVO]" in message:
            match = re.search(r"PORTAL: (\w+)", message, re.IGNORECASE)
            if match:
                self.current_portal = match.group(1).lower()
                self.intervention = None
                socketio.emit('portal_change', {"portal": self.current_portal}, namespace='/bot')

        # Actualizar contadores y emitir eventos específicos
        msg = message.strip()
        portal = self.current_portal or ""

        # --- Postulación exitosa ---
        if ("[ÉXITO]" in message or "[EXITO]" in message
                or "Postulación completada" in message
                or "-> ✅ Postulado" in message):
            with self.lock:
                self.stats["applied"] += 1
                self.stats["total"]   += 1
            title_m = re.search(r"(?:Postulación completada para|para):\s*(.+)", message)
            title = title_m.group(1).strip() if title_m else msg[:60]
            socketio.emit('update_stats', self.stats, namespace='/bot')
            socketio.emit('job_applied', {"portal": portal, "title": title, "status": "success"}, namespace='/bot')

        # --- Postulación externa registrada ---
        elif ("external_apply" in message or "-> ✅ external:" in message
              or ("[OK]" in message and "external:" in message)):
            with self.lock:
                self.stats["external"] += 1
                self.stats["total"]    += 1
            socketio.emit('update_stats', self.stats, namespace='/bot')

        # --- Sin navegación / no_navigation ---
        elif ("no_navigation" in message or "sin navegación" in message
              or "no navigation" in message.lower()):
            with self.lock:
                self.stats["no_nav"] += 1
                self.stats["total"]  += 1
            socketio.emit('update_stats', self.stats, namespace='/bot')

        # --- Filtrada / descartada ---
        elif ("[FILTRO]" in message or "skipped" in message.lower()
              or "Descartad" in message or "OFERTA_CERRADA" in message
              or "URL_MUERTA" in message or "filtrada" in message.lower()):
            with self.lock:
                self.stats["filtered"] += 1
                self.stats["total"]    += 1
            socketio.emit('update_stats', self.stats, namespace='/bot')

        # --- Error real ---
        elif ("[FALLO]" in message or "-> ❌" in message
              or re.search(r"\[OK\].*error:", message, re.IGNORECASE)):
            with self.lock:
                self.stats["errors"] += 1
                self.stats["total"]  += 1
            socketio.emit('update_stats', self.stats, namespace='/bot')
            socketio.emit('job_applied', {"portal": portal, "title": msg[:60], "status": "error"}, namespace='/bot')
        elif "[CAPTCHA]" in message or "CAPTCHA DETECTADO" in message.upper():
            portal = self._portal_from_message(message) or self.current_portal or ""
            with self.lock:
                self.intervention = {"type": "captcha", "portal": portal, "message": message.strip()}
            socketio.emit('captcha_required', {"message": "Verificación humana requerida en el navegador", "portal": portal}, namespace='/bot')
        elif "[SESION_NUEVA]" in message:
            # Navegador abierto para login manual antes de la sesion headless
            portal = self._portal_from_message(message) or self.current_portal or ""
            with self.lock:
                self.intervention = {"type": "login", "portal": portal, "message": message.strip()}
            socketio.emit('login_required', {
                "portal":  portal,
                "message": message.strip(),
                "manual":  True,
            }, namespace='/bot')
        elif "[SESION_INICIADA]" in message:
            with self.lock:
                self.intervention = None
            socketio.emit('login_resolved', {"message": message.strip()}, namespace='/bot')
            socketio.emit('session_status', get_session_status(), namespace='/bot')
        elif "[LOGIN_REQUERIDO]" in message:
            portal = self._portal_from_message(message) or self.current_portal or ""
            with self.lock:
                self.intervention = {"type": "login", "portal": portal, "message": message.strip()}
            socketio.emit('login_required', {"portal": portal, "message": message.strip()}, namespace='/bot')
        elif "[PREGUNTA_PENDIENTE]" in message:
            # Extraer texto de la pregunta del log y emitir evento al dashboard
            q_match = re.search(r'\[PREGUNTA_PENDIENTE\]\s*(.+)', message)
            question_text = q_match.group(1).strip() if q_match else message.strip()
            socketio.emit('pending_question', {"question": question_text}, namespace='/bot')

        elif "[BUSQUEDA]" in message:
            kw_match = re.search(r"Buscando: '(.+)' en", message)
            if kw_match:
                socketio.emit('keyword_update', {"keyword": kw_match.group(1)}, namespace='/bot')
        elif "[PROGRESO]" in message:
            prog_match = re.search(r"Aplicadas (\d+)/(\d+) en (\w+)", message)
            if prog_match:
                socketio.emit('portal_progress', {
                    "portal":   prog_match.group(3).lower(),
                    "applied":  int(prog_match.group(1)),
                    "max":      int(prog_match.group(2)),
                    "finished": False,
                }, namespace='/bot')
        elif "[PROGRESO_FINAL]" in message:
            prog_match = re.search(r"Aplicadas (\d+)/(\d+) en (\w+)", message)
            if prog_match:
                socketio.emit('portal_progress', {
                    "portal":   prog_match.group(3).lower(),
                    "applied":  int(prog_match.group(1)),
                    "max":      int(prog_match.group(2)),
                    "finished": True,
                }, namespace='/bot')

    def _portal_from_message(self, message):
        text = message.lower()
        for portal in _KNOWN_PORTALS:
            if portal in text:
                return portal
        return None

    def clear_logs(self):
        with self.lock:
            self.logs = []
            self.stats = {"applied": 0, "external": 0, "filtered": 0, "errors": 0, "no_nav": 0, "total": 0}
        socketio.emit('update_stats', self.stats, namespace='/bot')

    def set_process(self, proc):
        """Compatibilidad legado — asigna al proceso de apply."""
        with self.lock:
            self.apply_process = proc

    def set_scan_process(self, proc):
        with self.lock:
            self.scan_process = proc

    def set_apply_process(self, proc):
        with self.lock:
            self.apply_process = proc

    def get_status(self):
        with self.lock:
            scan_live  = self.scan_process  is not None and self.scan_process.poll()  is None
            apply_live = self.apply_process is not None and self.apply_process.poll() is None
            return {
                "running":       self.is_active,
                "scan_active":   self.scan_active,
                "apply_active":  self.apply_active,
                "process_active": scan_live or apply_live,
                "logs":          list(self.logs),
                "stats":         dict(self.stats),
                "current_portal": self.current_portal,
                "intervention":  self.intervention,
                "run_id":        self.scan_run_id + self.apply_run_id,
            }

    def _kill_proc(self, proc):
        """
        Mata el proceso y todos sus hijos.
        1. Escribe archivo señal → engine.py cierra browser Playwright limpiamente
        2. Espera hasta 2s para shutdown limpio
        3. Si aún vive: taskkill /F /T + terminate + kill (forzado)
        4. Limpia el archivo señal
        """
        _write_stop_signal()
        pid = getattr(proc, 'pid', None)

        # Dar 2 segundos para que engine.py detecte la señal y cierre solo
        import time as _t
        for _ in range(10):            # 10 x 200ms = 2s
            if proc.poll() is not None:
                break
            _t.sleep(0.2)

        # Si todavía vive → kill forzado
        if proc.poll() is None:
            if pid:
                try:
                    subprocess.run(
                        ["taskkill", "/F", "/T", "/PID", str(pid)],
                        capture_output=True, timeout=5
                    )
                except Exception:
                    pass
            try:
                proc.terminate()
            except Exception:
                pass
            try:
                _t.sleep(0.3)
                if proc.poll() is None:
                    proc.kill()
            except Exception:
                pass

        _clear_stop_signal()

    def stop_process(self):
        procs = []
        with self.lock:
            self.stop_requested = True
            for p in [self.scan_process, self.apply_process]:
                if p and p.poll() is None:
                    procs.append(p)
            self.scan_process  = None
            self.apply_process = None
            self.scan_active   = False
            self.apply_active  = False
            self.intervention  = None
        for p in procs:
            self._kill_proc(p)
        return bool(procs)

    # ── Scan ──────────────────────────────────────────────────────────────────
    def start_scan(self):
        with self.lock:
            self.scan_run_id += 1
            self.scan_active = True
            self.stop_requested = False
            run_id = self.scan_run_id
        return run_id

    def finish_scan(self, run_id):
        with self.lock:
            if run_id != self.scan_run_id:
                return False
            self.scan_active  = False
            self.scan_process = None
            return True

    # ── Apply ─────────────────────────────────────────────────────────────────
    def start_apply(self):
        with self.lock:
            self.apply_run_id += 1
            self.apply_active = True
            self.stop_requested = False
            self.current_portal = None
            self.intervention = None
            run_id = self.apply_run_id
        self.clear_logs()
        return run_id

    def finish_apply(self, run_id):
        with self.lock:
            if run_id != self.apply_run_id:
                return False
            self.apply_active  = False
            self.apply_process = None
            self.current_portal = None
            self.intervention  = None
            return True

    # ── Legado (usado por run_bot_thread) ─────────────────────────────────────
    def start_run(self):
        return self.start_apply()

    def finish_run(self, run_id):
        return self.finish_apply(run_id)

state = BotState()

_SESSIONS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sessions')
_KNOWN_PORTALS = [
    # Portales Chile
    'chiletrabajos', 'laborum', 'getonyboard', 'computrabajo', 'linkedin',
    'trabajando', 'infojobs',
    # Portales remotos internacionales (sin login, postulación externa)
    'weworkremotely', 'remotive', 'remoteco',
]


def _validate_portals(raw) -> list:
    """
    Filtra una lista de portales recibida del cliente contra la whitelist.
    Retorna solo los portales válidos. Nunca lanza excepción.
    """
    if not isinstance(raw, list):
        return []
    return [p for p in raw if isinstance(p, str) and p in _KNOWN_PORTALS]


# Indeed excluido: bloqueado por Cloudflare Turnstile
_PERSISTED_ENV_KEYS = {
    'USER_KEYWORDS',
    'USER_MAX_OFFERS',
    'USER_CV_PATH',
    'USER_FULL_NAME',
    'USER_FIRST_NAME',
    'USER_LAST_NAME',
    'USER_EMAIL',
    'USER_PHONE',
    'USER_PHONE_NUMBER',
    'USER_COUNTRY_CODE',
    'USER_COUNTRY',
    'USER_CITY',
    'USER_LINKEDIN',
    'USER_PORTFOLIO',
    'USER_SALARY',
    'USER_YEARS_EXP',
    'USER_AVAILABILITY',
    'USER_ENGLISH_LEVEL',
    'USER_WORK_MODE',
    'USER_COVER_LETTER',
    'LABORUM_EMAIL',
    'LABORUM_PASSWORD',
}
# SECURITY: keys que NUNCA se devuelven al browser vía /api/config
_SECRET_ENV_KEYS = {'LABORUM_PASSWORD', 'SECRET_KEY', 'SMTP_PASS'}
# Keys públicas = persistidas - secretas
_PUBLIC_ENV_KEYS = _PERSISTED_ENV_KEYS - _SECRET_ENV_KEYS
# Campos que se pasan al proceso del bot como env vars en tiempo de ejecución
_SENSITIVE_ENV_KEYS = _PERSISTED_ENV_KEYS


def update_env_values(env_path: str, updates: dict, remove_keys=None) -> None:
    """Actualiza .env sin reemplazar el archivo por un temporal.

    Thread-safe: usa _env_write_lock para evitar corrupción por escrituras concurrentes.
    En Windows + OneDrive, el rename atómico que usa python-dotenv puede fallar
    con PermissionError aunque el archivo sea escribible.
    """
    with _env_write_lock:
        existing_lines = []
        seen = set()
        remove_keys = set(remove_keys or [])

        if os.path.exists(env_path):
            with open(env_path, 'r', encoding='utf-8') as f:
                existing_lines = f.readlines()

        with open(env_path, 'w', encoding='utf-8', newline='') as f:
            for line in existing_lines:
                stripped = line.strip()
                if not stripped or stripped.startswith('#') or '=' not in line:
                    f.write(line)
                    continue

                key, _, _ = line.partition('=')
                key = key.strip()
                if key in remove_keys:
                    continue
                if key in updates:
                    f.write(f"{key}={updates[key]}\n")
                    seen.add(key)
                else:
                    f.write(line)

            for key, value in updates.items():
                if key not in seen:
                    f.write(f"{key}={value}\n")


def clean_form_value(value: str) -> str:
    value = str(value or '').strip()
    # Solo quitar comillas al inicio/final — nunca tocar barras invertidas (rutas Windows)
    value = value.strip("'\"")
    return re.sub(r'\s+', ' ', value).strip()


_COOKIES_MIN_BYTES = 8_000    # mismo umbral que engine.py (~8 KB)

def _portal_cookies_ok(portal: str) -> bool:
    """True si el archivo Cookies del portal tiene tamaño suficiente."""
    cookies_path = os.path.join(_SESSIONS_DIR, portal, "Default", "Network", "Cookies")
    try:
        return os.path.getsize(cookies_path) >= _COOKIES_MIN_BYTES
    except OSError:
        return False


def get_session_status() -> dict:
    """Detecta qué portales tienen sesión válida (archivo Cookies >= 35 KB)."""
    status = {}
    for portal in _KNOWN_PORTALS:
        status[portal] = _portal_cookies_ok(portal)
    return status


# Portales que requieren login real para postular
_PORTALS_REQUIRE_LOGIN = {
    'linkedin', 'computrabajo', 'laborum', 'trabajando',
    'infojobs', 'chiletrabajos', 'getonyboard',
}

# Señales DOM de sesión activa por portal (espejo del engine)
_LOGIN_SIGNALS_SERVER = {
    "linkedin":      ["div.global-nav__me-photo", "img.global-nav__me-photo",
                      "a[href*='/in/']", "[data-control-name='nav.settings']"],
    "computrabajo":  ["a[href*='/candidato/']", "a[href*='mi-cv']",
                      "[class*='user-name']", "span[class*='userName']"],
    "laborum":       ["a[href*='/postulantes/']", "[class*='userAvatar']",
                      "a:has-text('Mi cuenta')", "img[alt*='avatar' i]"],
    "trabajando":    ["a[href*='/mi-cv']", "a[href*='/mi-cuenta']",
                      "[class*='user-menu']", "[class*='userMenu']"],
    "infojobs":      ["a[href*='/candidato/mis-candidaturas']",
                      "a[href*='/candidato/mi-perfil']",
                      "[data-testid='user-menu']", "[class*='UserMenu']"],
    "chiletrabajos": ["a[href*='/mi-cuenta']", "a[href*='/mis-postulaciones']",
                      "[class*='user-logged']"],
    "getonyboard":   ["a[href*='/postulantes/']", "a[href*='/dashboard']",
                      "[class*='user-avatar']", "img[alt*='avatar' i]"],
}

_HOME_URLS_SERVER = {
    "linkedin":      "https://www.linkedin.com/feed",
    "computrabajo":  "https://cl.computrabajo.com/candidato/",
    "laborum":       "https://www.laborum.cl/postulantes/dashboard",
    "trabajando":    "https://www.trabajando.cl/mi-cuenta",
    "infojobs":      "https://www.infojobs.net/candidato/mis-candidaturas",
    "chiletrabajos": "https://www.chiletrabajos.cl/mi-cuenta",
    "getonyboard":   "https://www.getonbrd.com/postulantes/dashboard",
}

_session_verify_lock = threading.Lock()
_session_verify_running = False


_LOGIN_PAGE_SIGNALS = [
    "input[type='password']",
    "form[action*='login']",
    "form[action*='signin']",
    "a:has-text('Iniciar sesión')",
    "button:has-text('Iniciar sesión')",
    "a:has-text('Sign in')",
    "button:has-text('Sign In')",
    "input[name='session_password']",
    "input[id*='password']",
]

_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)


def _verify_session_headless(portal: str) -> str:
    """
    Abre browser headless con user-agent real, navega a la home del portal
    y verifica sesión por DOM + URL. Robusto ante SPAs y anti-bot básico.
    Retorna: 'ok' | 'expired' | 'no_cookies' | 'error'
    """
    if not _portal_cookies_ok(portal):
        return "no_cookies"

    session_dir = os.path.join(_SESSIONS_DIR, portal)
    home_url    = _HOME_URLS_SERVER.get(portal)
    sels        = _LOGIN_SIGNALS_SERVER.get(portal, [])

    if not home_url:
        return "ok"   # sin URL de check → asumir ok si cookies existen

    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            ctx = pw.chromium.launch_persistent_context(
                session_dir,
                headless    = True,
                user_agent  = _UA,
                viewport    = {"width": 1280, "height": 800},
                locale      = "es-CL",
                timezone_id = "America/Santiago",
                args        = [
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                ],
                ignore_default_args=["--enable-automation"],
            )
            pg = ctx.new_page()
            # Ocultar webdriver
            pg.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
            )
            try:
                # Navegar — esperar hasta networkidle para SPAs
                try:
                    pg.goto(home_url, wait_until="networkidle", timeout=25_000)
                except Exception:
                    try:
                        pg.goto(home_url, wait_until="domcontentloaded", timeout=20_000)
                        pg.wait_for_timeout(4000)
                    except Exception:
                        pass

                pg.wait_for_timeout(2500)   # margen extra para render JS

                current_url = pg.url.lower()

                # Si fue redirigido a una página de login → expirada
                login_keywords = ["login", "signin", "sign-in", "account/login",
                                  "candidato/login", "iniciar-sesion", "auth"]
                if any(k in current_url for k in login_keywords):
                    return "expired"

                # Chequear señales negativas (form de login visible)
                for bad_sel in _LOGIN_PAGE_SIGNALS:
                    try:
                        el = pg.query_selector(bad_sel)
                        if el and el.is_visible():
                            return "expired"
                    except Exception:
                        pass

                # Chequear señales positivas de sesión activa
                if sels:
                    for sel in sels:
                        try:
                            el = pg.query_selector(sel)
                            if el and el.is_visible():
                                return "ok"
                        except Exception:
                            pass

                    # Segunda vuelta con wait_for_selector (más lenta pero fiable)
                    for sel in sels[:3]:
                        try:
                            pg.wait_for_selector(sel, timeout=4_000)
                            return "ok"
                        except Exception:
                            pass

                    return "expired"
                else:
                    # Sin selectores definidos → si URL no redirigió a login, asumir ok
                    return "ok"

            finally:
                try:
                    ctx.close()
                except Exception:
                    pass

    except Exception as exc:
        _log.warning("[CHECK_SESSION] Error verificando %s: %s", portal, exc)
        return "error"


def run_bot_thread(portals, runtime_env=None):
    run_id = state.start_run()
    
    socketio.emit('bot_status', state.get_status() | {"status": "started"}, namespace='/bot')
    
    try:
        # Unificar portales en una sola llamada para máxima velocidad
        portal_str = ",".join(portals)
        state.add_log(f"\n[SISTEMA] Iniciando ejecución unificada: {portal_str}\n")
        
        cmd = [sys.executable, "-u", "main.py", "--portal", portal_str, "--multi-keyword"]
        child_env = os.environ.copy()
        child_env['PYTHONUTF8']       = '1'
        child_env['PYTHONIOENCODING'] = 'utf-8'
        child_env['PYTHONLEGACYWINDOWSSTDIO'] = '0'
        if runtime_env:
            child_env.update(runtime_env)

        try:
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                errors='replace',
                bufsize=1,
                env=child_env
            )
            state.set_process(process)
            _stream_process(process, "", lambda: None)
        except Exception as e:
            state.add_log(f"\n[FALLO] Error crítico en ejecución: {str(e)}\n")
    finally:
        state.add_log("\n[SISTEMA] Ejecución maestra finalizada.\n")
        if state.finish_run(run_id):
            socketio.emit('bot_status', state.get_status() | {"status": "finished"}, namespace='/bot')
            socketio.emit('session_status', get_session_status(), namespace='/bot')

@app.route('/')
def index():
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/bot-state')
def api_bot_state():
    """Estado mínimo del bot para sincronización al reconectar el socket."""
    scan_live  = state.scan_process  is not None and state.scan_process.poll()  is None
    apply_live = state.apply_process is not None and state.apply_process.poll() is None
    return jsonify({
        "running":       state.is_active,
        "scan_active":   state.scan_active,
        "apply_active":  state.apply_active,
        "process_active": scan_live or apply_live,
    })


def _resolve_portals(data: dict) -> str:
    """Convierte el campo 'portals' (array) o 'portal' (string) en un string
    separado por comas para pasarlo como --portal a main.py.
    Si no se indica ninguno, retorna None (main.py usará _ALL_PORTALS).
    SECURITY: valida cada nombre contra la whitelist de portales conocidos."""
    _VALID = set(_KNOWN_PORTALS) | {'indeed'}
    portals = data.get('portals')  # array del frontend
    if portals and isinstance(portals, list) and portals:
        safe = [p.strip().lower() for p in portals if p.strip().lower() in _VALID]
        return ','.join(safe) if safe else None
    single = data.get('portal', '').strip().lower()
    return single if single in _VALID else None


def _make_child_env(extra=None):
    e = os.environ.copy()
    e.update({'PYTHONUTF8': '1', 'PYTHONIOENCODING': 'utf-8', 'PYTHONLEGACYWINDOWSSTDIO': '0'})
    if extra:
        e.update({k: v for k, v in extra.items() if v})
    return e

def _stream_process(proc, finish_log: str, finish_cb, stop_check=None):
    """
    Lee stdout sin bloquear el thread principal.
    Usa un reader-thread + queue para que el check de stop_requested
    corra cada 200ms sin importar si el proceso está imprimiendo o no.
    """
    import queue as _queue

    q = _queue.Queue()

    def _reader():
        try:
            for line in iter(proc.stdout.readline, ""):
                q.put(line)
        except Exception:
            pass
        finally:
            q.put(None)          # sentinel — proceso terminó

    threading.Thread(target=_reader, daemon=True).start()

    _stop = stop_check or (lambda: state.stop_requested)

    while True:
        if _stop():
            state._kill_proc(proc)
            # Vaciar la queue para no dejar el reader-thread colgado
            while True:
                try: q.get_nowait()
                except Exception: break
            break
        try:
            line = q.get(timeout=0.2)
            if line is None:        # proceso terminó normalmente
                break
            if line:
                state.add_log(line)
        except Exception:
            if proc.poll() is not None:   # proceso ya murió
                # Drenar lo que quede
                while True:
                    try:
                        line = q.get_nowait()
                        if line and line is not None:
                            state.add_log(line)
                    except Exception:
                        break
                break

    try:
        proc.stdout.close()
    except Exception:
        pass
    try:
        proc.wait(timeout=3)
    except Exception:
        pass
    # Limpiar procesos Chromium zombie que el subprocess pudo haber dejado
    if proc.pid:
        _kill_chromium_children(proc.pid)
    state.add_log(finish_log)
    finish_cb()


@app.route('/api/scan', methods=['POST'])
def api_scan():
    """Escanea ofertas sin postular — corre independiente del proceso de apply."""
    if state.scan_active:
        return jsonify({'ok': False, 'msg': 'Ya hay un scan en curso.'})
    data    = request.json or {}
    portals = _resolve_portals(data)
    label   = portals or 'todos los portales'

    def _run():
        run_id = state.start_scan()
        socketio.emit('bot_status', state.get_status() | {"status": "scan_started"}, namespace='/bot')
        try:
            cmd = [sys.executable, "-u", "main.py", "--scan"]
            if portals:
                cmd += ["--portal", portals]
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, encoding='utf-8', errors='replace', bufsize=1,
                                    env=_make_child_env())
            state.set_scan_process(proc)
            _start_watchdog(proc, PORTAL_TIMEOUT_S, f"scan-{portals or 'all'}")
            def _finish():
                if state.finish_scan(run_id):
                    socketio.emit('bot_status', state.get_status() | {"status": "scan_finished"}, namespace='/bot')
            _stream_process(proc, "\n[SCAN] Escaneo completado.\n", _finish)
        except Exception as e:
            state.add_log(f"\n[SCAN] Error: {e}\n")
            state.finish_scan(run_id)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'ok': True, 'msg': f'Scan iniciado para {label}'})


@app.route('/api/postular', methods=['POST'])
def api_postular():
    """Postulación directa — recorre portales y aplica sin necesitar scan previo."""
    if state.apply_active:
        return jsonify({'ok': False, 'msg': 'Ya hay una postulación en curso.'})
    data    = request.json or {}
    portals = _resolve_portals(data)
    label   = portals or 'todos los portales'
    runtime_env = {}
    if data.get('keywords'):
        runtime_env['USER_KEYWORDS'] = str(data['keywords']).strip().strip("'\"")[:200]
    if data.get('max_offers'):
        # SECURITY: validar que sea entero en rango [1, 100]
        try:
            _max = max(1, min(100, int(str(data['max_offers']).strip())))
            runtime_env['USER_MAX_OFFERS'] = str(_max)
        except (ValueError, TypeError):
            pass  # ignorar valor inválido, usar el default del .env

    def _run():
        run_id = state.start_apply()
        socketio.emit('bot_status', state.get_status() | {"status": "started"}, namespace='/bot')
        try:
            cmd = [sys.executable, "-u", "main.py", "--multi-keyword"]
            if portals:
                cmd += ["--portal", portals]
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, encoding='utf-8', errors='replace', bufsize=1,
                                    env=_make_child_env(runtime_env))
            state.set_apply_process(proc)
            _start_watchdog(proc, MASTER_TIMEOUT_S, f"postular-{portals or 'all'}")
            def _finish():
                if state.finish_apply(run_id):
                    socketio.emit('bot_status', state.get_status() | {"status": "finished"}, namespace='/bot')
            _stream_process(proc, "\n[POSTULAR] Postulación completada.\n", _finish)
        except Exception as e:
            state.add_log(f"\n[POSTULAR] Error: {e}\n")
            state.finish_apply(run_id)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'ok': True, 'msg': f'Postulación iniciada para {label}'})


@app.route('/api/apply_queue', methods=['POST'])
def api_apply_queue():
    """Aplica a ofertas en cola (scan previo). Independiente del scan."""
    if state.apply_active:
        return jsonify({'ok': False, 'msg': 'Ya hay una postulación en curso.'})
    data    = request.json or {}
    portals = _resolve_portals(data)
    label   = portals or 'todos los portales'

    def _run():
        run_id = state.start_apply()
        socketio.emit('bot_status', state.get_status() | {"status": "started"}, namespace='/bot')
        try:
            cmd = [sys.executable, "-u", "main.py", "--apply-queue"]
            if portals:
                cmd += ["--portal", portals]
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, encoding='utf-8', errors='replace', bufsize=1,
                                    env=_make_child_env())
            state.set_apply_process(proc)
            _start_watchdog(proc, PORTAL_TIMEOUT_S, f"apply-queue-{portals or 'all'}")
            def _finish():
                if state.finish_apply(run_id):
                    socketio.emit('bot_status', state.get_status() | {"status": "finished"}, namespace='/bot')
            _stream_process(proc, "\n[APPLY-QUEUE] Completado.\n", _finish)
        except Exception as e:
            state.add_log(f"\n[APPLY-QUEUE] Error: {e}\n")
            state.finish_apply(run_id)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'ok': True, 'msg': f'Apply-queue iniciado para {label}'})


@app.route('/api/keywords')
def api_keywords():
    """Devuelve KEYWORD_GROUPS desde config.py para que el dashboard los muestre como chips."""
    from bot.config import KEYWORD_GROUPS
    return jsonify(KEYWORD_GROUPS)

@app.route('/api/config')
def api_config():
    """Devuelve las variables del .env como JSON para que el dashboard las cargue vía fetch.
    SECURITY: solo devuelve _PUBLIC_ENV_KEYS — nunca contraseñas ni SECRET_KEY."""
    env_data = {}
    if os.path.exists('.env'):
        with open('.env', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    key, _, val = line.partition('=')
                    key = key.strip()
                    if key in _PUBLIC_ENV_KEYS:
                        env_data[key] = val.strip()
    return jsonify(env_data)

@app.route('/api/parse_cv', methods=['POST'])
def api_parse_cv():
    """Recibe un CV, lo guarda en uploads/, extrae campos y actualiza USER_CV_PATH en .env."""
    tmp_path = None
    try:
        cv_file = request.files.get('cv_file')
        if not cv_file or cv_file.filename == '':
            return jsonify({"status": "error", "message": "No se recibió archivo"})

        ext = os.path.splitext(cv_file.filename)[1].lower()
        if ext not in ('.pdf', '.docx', '.doc'):
            return jsonify({"status": "error", "message": f"Formato no soportado: {ext}. Usa PDF o DOCX."})

        # Guardar permanentemente en uploads/ (el bot lo usa al postular)
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        safe_name = "cv" + ext
        permanent_path = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], safe_name))
        cv_file.save(permanent_path)

        # Parsear el CV desde la copia permanente
        from bot.cv_parser import parse_cv
        fields = parse_cv(permanent_path)

        # Persistir la ruta en .env para que el bot siempre la encuentre
        env_path = os.path.abspath('.env')
        if not os.path.exists(env_path):
            open(env_path, 'w', encoding='utf-8').close()
        update_env_values(env_path, {'USER_CV_PATH': permanent_path})

        return jsonify({
            "status": "success",
            "fields": fields,
            "filename": cv_file.filename,
            "cv_path": os.path.basename(permanent_path),  # SECURITY: solo basename, no ruta absoluta
        })
    except Exception as e:
        import logging as _log; _log.getLogger("applyjob").error("parse_cv error: %s", e)
        return jsonify({"status": "error", "message": "Error al procesar el CV. Intenta de nuevo."})


@app.route('/save_config', methods=['POST'])
def save_config():
    try:
        _config_rate_limiter.check()
    except RuntimeError as e:
        return jsonify({'ok': False, 'error': str(e)}), 429
    try:
        data = request.form.to_dict()
        
        env_path = os.path.abspath('.env')
        if not os.path.exists(env_path):
            with open(env_path, 'w', encoding='utf-8') as f: f.write("")

        updates = {}

        if 'MAX_OFFERS' in data:
            updates['USER_MAX_OFFERS'] = clean_form_value(data.pop('MAX_OFFERS'))

        for key, val in data.items():
            if key in _PERSISTED_ENV_KEYS and val.strip():
                updates[key] = clean_form_value(val)

        # Guardar sin borrar ningún campo — todos se persisten en .env
        update_env_values(env_path, updates)

        return jsonify({"status": "success", "message": "Configuracion guardada correctamente."})
    except Exception as e:
        import logging as _log; _log.getLogger("applyjob").error("save_config error: %s", e)
        return jsonify({"status": "error", "message": "Error al guardar configuración."})

@app.route('/api/stats')
def api_stats():
    """
    Devuelve estadísticas globales y por portal desde la DB SQLite.
    El dashboard las carga al iniciar para mostrar el historial acumulado.
    """
    try:
        from bot.state import get_stats
        stats = get_stats()
        by_portal = stats.get("by_portal", {})

        # Agregar by_status desde by_portal (get_stats no lo devuelve directamente)
        by_status: dict = {}
        for portal_data in by_portal.values():
            for status, cnt in portal_data.items():
                by_status[status] = by_status.get(status, 0) + cnt

        applied = by_status.get("applied", 0)
        errors  = sum(v for k, v in by_status.items() if "error" in k.lower())
        skipped = sum(v for k, v in by_status.items()
                      if k.startswith("skipped") or k.startswith("external") or k == "dry_run")

        return jsonify({
            "total":     stats.get("total", 0),
            "applied":   applied,
            "errors":    errors,
            "skipped":   skipped,
            "by_portal": by_portal,
            "by_status": by_status,
        })
    except Exception as e:
        return jsonify({"total": 0, "applied": 0, "errors": 0, "skipped": 0,
                        "by_portal": {}, "by_status": {}, "error": str(e)})


_RESTRICTIONS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'portal_restrictions.json')

@app.route('/api/check-sessions', methods=['POST'])
def api_check_sessions():
    """
    Verifica sesiones de portales via browser headless.
    Body JSON: { "portals": ["linkedin", "laborum"] }  — vacío = todos los que requieren login.
    Retorna: { portal: "ok"|"expired"|"no_cookies"|"error" }
    """
    global _session_verify_running
    with _session_verify_lock:
        if _session_verify_running:
            return jsonify({"error": "Verificación ya en curso"}), 429
        _session_verify_running = True

    try:
        requested = (request.json or {}).get("portals") or list(_PORTALS_REQUIRE_LOGIN)
        portals_to_check = [p for p in requested if p in _PORTALS_REQUIRE_LOGIN]

        results = {}
        # Primero chequeo rápido por cookies (sin browser)
        for portal in portals_to_check:
            results[portal] = "no_cookies" if not _portal_cookies_ok(portal) else "checking"

        # Emit progreso
        socketio.emit('session_check_progress', results, namespace='/bot')

        # Verificación DOM por browser para los que tienen cookies
        for portal in portals_to_check:
            if results[portal] == "checking":
                result = _verify_session_headless(portal)
                results[portal] = result
                socketio.emit('session_check_progress',
                              {portal: result}, namespace='/bot')

        # Actualizar dots del dashboard
        socketio.emit('session_status', get_session_status(), namespace='/bot')
        return jsonify(results)
    finally:
        with _session_verify_lock:
            _session_verify_running = False


@app.route('/api/portal-restrictions')
def api_portal_restrictions():
    """Devuelve el estado de restricciones por portal."""
    try:
        if not os.path.exists(_RESTRICTIONS_PATH):
            return jsonify({})
        with open(_RESTRICTIONS_PATH, encoding='utf-8') as f:
            return jsonify(json.load(f))
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route('/api/portal-restrictions/clear', methods=['POST'])
def api_clear_restriction():
    """Limpia la restricción de un portal específico."""
    portal = (request.json or {}).get('portal', '').strip().lower()
    if not portal:
        return jsonify({'ok': False, 'error': 'portal requerido'}), 400
    try:
        data = {}
        if os.path.exists(_RESTRICTIONS_PATH):
            with open(_RESTRICTIONS_PATH, encoding='utf-8') as f:
                data = json.load(f)
        if portal in data:
            data[portal]['restricted'] = False
            with open(_RESTRICTIONS_PATH, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/recent')
def api_recent():
    """Devuelve las últimas N postulaciones desde la DB SQLite."""
    try:
        limit = min(int(request.args.get('limit', 30)), 500)  # SECURITY: cap máximo
        from bot.state import get_recent
        rows = get_recent(limit)
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db-stats')
def api_db_stats():
    """Estadísticas acumuladas de la DB (totales históricos)."""
    try:
        from bot.state import get_stats
        return jsonify(get_stats())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


_PENDING_Q_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'pending_questions.json')
_QA_PATH          = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'question_answers.json')
_QA_CACHE_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'qa_cache.json')
_QUICK_LINKS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'quick_links.json')
_PROFILE_KB_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'profile_kb.json')
_ENV_PATH         = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')


# ---------------------------------------------------------------------------
# Quick Links API — ofertas de bodega/operario para postulación manual
# ---------------------------------------------------------------------------

@app.route('/api/quick-links')
def api_quick_links():
    """Devuelve los links rápidos de bodega pendientes (no descartados)."""
    try:
        if not os.path.exists(_QUICK_LINKS_PATH):
            return jsonify([])
        with open(_QUICK_LINKS_PATH, encoding='utf-8') as f:
            links = json.load(f)
        active = [l for l in links if not l.get('dismissed')]
        return jsonify(active)
    except Exception as e:
        return jsonify([])


@app.route('/api/quick-links/dismiss', methods=['POST'])
def api_quick_links_dismiss():
    """Marca un link como descartado (ya postulado o no interesa)."""
    data = request.json or {}
    url  = data.get('url', '').strip()
    if not url:
        return jsonify({'ok': False, 'error': 'url requerida'}), 400
    try:
        if not os.path.exists(_QUICK_LINKS_PATH):
            return jsonify({'ok': True})
        with open(_QUICK_LINKS_PATH, encoding='utf-8') as f:
            links = json.load(f)
        for l in links:
            if l.get('url') == url:
                l['dismissed'] = True
        with open(_QUICK_LINKS_PATH, 'w', encoding='utf-8') as f:
            json.dump(links, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/quick-links/clear', methods=['POST'])
def api_quick_links_clear():
    """Descarta todos los links activos."""
    try:
        if os.path.exists(_QUICK_LINKS_PATH):
            with open(_QUICK_LINKS_PATH, encoding='utf-8') as f:
                links = json.load(f)
            for l in links:
                l['dismissed'] = True
            with open(_QUICK_LINKS_PATH, 'w', encoding='utf-8') as f:
                json.dump(links, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/scan-quick-links', methods=['POST'])
def api_scan_quick_links():
    """
    Lanza run_scan_quick_links en un subproceso y devuelve los resultados.
    Escanea los quick_links.json guardados para recopilar preguntas de ATS externos.
    """
    import subprocess
    import threading

    result_holder = {}
    done_event = threading.Event()

    def _run():
        try:
            proc = subprocess.run(
                [sys.executable, '-u', 'main.py', '--scan-quick-links', '--headless'],
                capture_output=True, text=True, encoding='utf-8', errors='replace',
                cwd=os.path.dirname(os.path.abspath(__file__)),
                timeout=300,
            )
            output = proc.stdout + proc.stderr
            # Parse top_questions from output
            top_qs = []
            already = queued = failed = scanned = 0
            for line in output.splitlines():
                if '  Escaneadas' in line:
                    try: scanned = int(line.split(':')[-1].strip())
                    except ValueError: pass
                elif 'Ya respondidas' in line:
                    try: already = int(line.split(':')[-1].strip())
                    except ValueError: pass
                elif 'En cola' in line:
                    try: queued = int(line.split(':')[-1].strip())
                    except ValueError: pass
                elif 'Sin formulario' in line:
                    try: failed = int(line.split(':')[-1].strip())
                    except ValueError: pass
                elif line.strip().startswith(tuple(str(i)+'.' for i in range(1,20))):
                    # Líneas de formato "  N. [Mx] Pregunta..."
                    import re as _re
                    m = _re.match(r'\s*\d+\.\s*(?:\[\d+x\]\s*)?(.+)', line)
                    if m:
                        top_qs.append(m.group(1).strip()[:80])
            result_holder.update({
                'scanned': scanned, 'already_answered': already,
                'queued': queued, 'failed': failed,
                'top_questions': top_qs[:15],
                'output': output[-2000:],
            })
        except subprocess.TimeoutExpired:
            result_holder.update({'error': 'Timeout (300s)', 'scanned': 0})
        except Exception as e:
            result_holder.update({'error': str(e), 'scanned': 0})
        finally:
            done_event.set()

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    done_event.wait(timeout=310)

    if 'error' in result_holder and 'scanned' not in result_holder:
        return jsonify({'ok': False, 'error': result_holder.get('error', 'unknown')}), 500
    return jsonify({**result_holder, 'ok': True})


@app.route('/api/auto-answer-pending', methods=['POST'])
def api_auto_answer_pending():
    """
    Recorre pending_questions.json y responde automáticamente las que tienen
    answered=False usando _auto_answer() + _match_qa() del CV del usuario.
    No requiere browser — es solo procesamiento Python.
    """
    try:
        pending_path = _PENDING_Q_PATH
        qa_cache_path = _QA_CACHE_PATH

        if not os.path.exists(pending_path):
            return jsonify({'ok': True, 'answered': 0, 'still_pending': 0})

        with open(pending_path, encoding='utf-8') as f:
            pending = json.load(f)

        # Importar helpers del form_filler
        import sys as _sys
        _bot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'bot')
        if _bot_dir not in _sys.path:
            _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

        from bot.form_filler import _match_qa, _auto_answer, _normalize
        from bot.config import USER_PROFILE

        # Cargar qa_cache para persistir respuestas nuevas
        qa_cache = {}
        if os.path.exists(qa_cache_path):
            try:
                with open(qa_cache_path, encoding='utf-8') as f:
                    qa_cache = json.load(f)
            except Exception:
                qa_cache = {}

        newly_answered = 0
        for entry in pending:
            if entry.get('answered'):
                continue  # ya tiene respuesta

            label = entry.get('label', '')
            norm  = entry.get('norm', '') or _normalize(label)
            if not label:
                continue

            # Intentar respuesta automática
            ans = _match_qa(label) or _auto_answer(label, USER_PROFILE)
            if not ans:
                # Fallback: cover_letter como último recurso para preguntas abiertas
                cl = USER_PROFILE.get('cover_letter', '').strip()
                if cl and any(kw in label.lower() for kw in (
                    'experiencia', 'presentate', 'cuéntanos', 'cuentanos',
                    'sobre ti', 'motivacion', 'motivación', 'background',
                    'habilidades', 'conocimientos', 'perfil', 'describe',
                )):
                    ans = cl

            if ans:
                entry['answered'] = True
                entry['answer']   = str(ans)
                entry['source']   = 'auto'
                # Guardar en qa_cache también
                if norm not in qa_cache:
                    qa_cache[norm] = str(ans)
                newly_answered += 1

        # Guardar cambios
        with open(pending_path, 'w', encoding='utf-8') as f:
            json.dump(pending, f, ensure_ascii=False, indent=2)
        with open(qa_cache_path, 'w', encoding='utf-8') as f:
            json.dump(qa_cache, f, ensure_ascii=False, indent=2)

        still_pending = sum(1 for e in pending if not e.get('answered'))
        return jsonify({
            'ok': True,
            'answered': newly_answered,
            'still_pending': still_pending,
            'total': len(pending),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/apply-quick-links', methods=['POST'])
def api_apply_quick_links():
    """
    Aplica automáticamente a los quick_links.json (ATS externos).
    Lanza run_apply_quick_links en subproceso — procesa hasta max (default 5).
    """
    import subprocess
    import threading as _th

    data     = request.json or {}
    max_apply = max(1, min(20, int(data.get('max', 5))))

    result_holder = {}
    done_event    = _th.Event()

    def _run():
        try:
            proc = subprocess.run(
                [sys.executable, '-u', 'main.py',
                 '--apply-quick-links', '--headless', '--max', str(max_apply)],
                capture_output=True, text=True, encoding='utf-8', errors='replace',
                cwd=os.path.dirname(os.path.abspath(__file__)),
                timeout=360,
            )
            output = proc.stdout + proc.stderr
            applied = pending = failed = 0
            for line in output.splitlines():
                if 'aplicadas' in line.lower():
                    import re as _re
                    m = _re.search(r'(\d+)\s+aplicadas', line, _re.I)
                    if m: applied = int(m.group(1))
                if 'pendientes' in line.lower():
                    import re as _re
                    m = _re.search(r'(\d+)\s+pendientes', line, _re.I)
                    if m: pending = int(m.group(1))
                if 'fallidas' in line.lower():
                    import re as _re
                    m = _re.search(r'(\d+)\s+fallidas', line, _re.I)
                    if m: failed = int(m.group(1))
            result_holder.update({
                'applied': applied, 'pending': pending, 'failed': failed,
                'output': output[-1500:],
            })
        except subprocess.TimeoutExpired:
            result_holder.update({'error': 'Timeout (360s)', 'applied': 0})
        except Exception as exc:
            result_holder.update({'error': str(exc), 'applied': 0})
        finally:
            done_event.set()

    t = _th.Thread(target=_run, daemon=True)
    t.start()
    done_event.wait(timeout=370)

    if 'error' in result_holder and 'applied' not in result_holder:
        return jsonify({'ok': False, 'error': result_holder.get('error')}), 500
    return jsonify({**result_holder, 'ok': True})


@app.route('/api/fill-curriculum', methods=['POST'])
def api_fill_curriculum():
    """
    Abre trabajando.cl/mi-curriculum#/ con la sesión guardada
    y rellena los campos del perfil con los datos del usuario.
    """
    result_holder = {}
    done_event = threading.Event()

    def _run():
        try:
            proc = subprocess.run(
                [sys.executable, '-u', 'main.py', '--fill-curriculum', '--headless'],
                capture_output=True, text=True, encoding='utf-8', errors='replace',
                cwd=os.path.dirname(os.path.abspath(__file__)),
                timeout=120,
            )
            output = proc.stdout + proc.stderr
            ok = 'curriculum guardado' in output.lower() or proc.returncode == 0
            result_holder.update({'ok': ok, 'output': output[-1500:]})
        except subprocess.TimeoutExpired:
            result_holder.update({'ok': False, 'error': 'Timeout (120s)'})
        except Exception as e:
            result_holder.update({'ok': False, 'error': str(e)})
        finally:
            done_event.set()

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    done_event.wait(timeout=130)
    return jsonify(result_holder)


def _normalize_srv(text: str) -> str:
    """Normaliza texto para comparaciones: minúsculas, sin tildes, sin espacios extra."""
    import unicodedata as _ud
    nfkd = _ud.normalize("NFKD", text.lower())
    ascii_t = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_t).strip()


def _load_merged_qa() -> dict:
    """Carga y fusiona question_answers.json + qa_cache.json (normalizado)."""
    result = {}
    for path in (_QA_PATH, _QA_CACHE_PATH):
        if not os.path.exists(path):
            continue
        try:
            with open(path, encoding='utf-8') as f:
                for k, v in json.load(f).items():
                    if not k or k.startswith('_') or k.startswith('─') or not v:
                        continue
                    nk = _normalize_srv(k)
                    if nk:
                        result[nk] = v
        except Exception:
            pass
    return result


def _load_profile_kb_qa() -> dict:
    """Carga los qa_overrides de profile_kb.json (todas las categorías)."""
    result = {}
    if not os.path.exists(_PROFILE_KB_PATH):
        return result
    try:
        with open(_PROFILE_KB_PATH, encoding='utf-8') as f:
            kb = json.load(f)
        for category in kb.values():
            for q, a in (category.get('qa_overrides') or {}).items():
                nk = _normalize_srv(q)
                if nk and a:
                    result[nk] = a
    except Exception:
        pass
    return result


def _load_profile_env_qa() -> dict:
    """
    Genera respuestas básicas desde las variables USER_* del .env.
    Cubre preguntas sobre nombre, email, teléfono, ciudad, salario, etc.
    """
    result: dict = {}
    cfg: dict = {}

    # Intentar leer .env
    if os.path.exists(_ENV_PATH):
        try:
            with open(_ENV_PATH, encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if '=' in line and not line.startswith('#'):
                        k, _, v = line.partition('=')
                        cfg[k.strip()] = v.strip().strip('"').strip("'")
        except Exception:
            pass

    # También leer desde config actual de la app
    try:
        from bot.config import USER_PROFILE
        cfg.update({k: str(v) for k, v in USER_PROFILE.items() if v})
    except Exception:
        pass

    # Mapeo: palabras clave de pregunta → valor del perfil
    name     = cfg.get('USER_FULL_NAME') or f"{cfg.get('USER_FIRST_NAME','')} {cfg.get('USER_LAST_NAME','')}".strip()
    email    = cfg.get('USER_EMAIL', '')
    phone    = cfg.get('USER_PHONE', cfg.get('USER_PHONE_NUMBER', ''))
    city     = cfg.get('USER_CITY', 'Maipú, Santiago')
    salary   = cfg.get('USER_SALARY', '850000')
    avail    = cfg.get('USER_AVAILABILITY', 'Inmediata')
    english  = cfg.get('USER_ENGLISH_LEVEL', 'Básico')
    wmode    = cfg.get('USER_WORK_MODE', 'Presencial')
    yexp     = cfg.get('USER_YEARS_EXP', '0')

    mappings = []
    if name:
        mappings += [
            ("nombre completo", name), ("your name", name),
            ("nombre y apellido", name), ("tu nombre", name),
        ]
    if email:
        mappings += [
            ("correo electronico", email), ("email", email),
            ("tu correo", email), ("email address", email),
        ]
    if phone:
        mappings += [
            ("telefono", phone), ("numero de telefono", phone),
            ("celular", phone), ("phone number", phone),
            ("numero celular", phone),
        ]
    if city:
        mappings += [
            ("ciudad de residencia", city), ("donde vives", city),
            ("ciudad", city), ("ubicacion", city),
        ]
    if salary:
        mappings += [
            ("pretension salarial", salary), ("pretension de renta", salary),
            ("renta esperada", salary), ("expectativa salarial", salary),
            ("cuanto quieres ganar", salary),
        ]
    if avail:
        mappings += [
            ("disponibilidad", avail), ("cuando puedes empezar", avail),
            ("fecha de incorporacion", avail), ("disponibilidad de incorporacion", avail),
        ]
    if english:
        mappings += [
            ("nivel de ingles", english), ("english level", english),
            ("hablas ingles", english),
        ]
    if wmode:
        mappings += [
            ("modalidad de trabajo", wmode), ("modalidad preferida", wmode),
            ("trabajo presencial o remoto", wmode),
        ]
    if yexp:
        mappings += [
            ("anos de experiencia", yexp), ("years of experience", yexp),
            ("cuantos anos de experiencia tienes", yexp),
        ]

    for q, a in mappings:
        nk = _normalize_srv(q)
        if nk:
            result[nk] = a

    return result


def _suggest_from_cache(norm: str) -> str:
    """
    Busca la mejor respuesta para una pregunta normalizada.
    Jerarquía:
      1. question_answers.json + qa_cache.json (exacto)
      2. profile_kb.json qa_overrides
      3. Variables USER_* del .env / config
      4. Substring match en todos los anteriores
      5. Word-overlap ≥ 60% en todos los anteriores
    """
    # Construir mapa unificado: QA principal > profile_kb > env
    qa = _load_merged_qa()
    kb_qa = _load_profile_kb_qa()
    env_qa = _load_profile_env_qa()

    # Fusionar (QA principal tiene prioridad)
    combined: dict = {}
    combined.update(env_qa)
    combined.update(kb_qa)
    combined.update(qa)  # máxima prioridad

    # 1. Match exacto
    if norm in combined:
        return combined[norm]

    # 2. Substring
    for k, v in combined.items():
        if len(k) >= 15 and k in norm:
            return v
        if len(norm) >= 15 and norm in k:
            return v

    # 3. Word-overlap ≥ 60%
    _STOP = {"de","la","el","en","y","a","con","su","tu","un","una","es","se",
             "si","no","para","que","por","al","del","lo","las","los","has",
             "have","your","you","the","and","or","is","are","do","did"}
    words_n = set(norm.split()) - _STOP
    best_v, best_s = None, 0.0
    for k, v in combined.items():
        words_k = set(k.split()) - _STOP
        if not words_n or not words_k:
            continue
        shared = words_n & words_k
        score = len(shared) / max(len(words_n), len(words_k))
        if score > best_s:
            best_s, best_v = score, v
    return best_v if best_s >= 0.55 else ""


_SCAN_QUEUE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'scan_queue.json')

@app.route('/api/scan-queue')
def api_scan_queue():
    """Devuelve todos los items de la cola de scan para mostrar en el panel."""
    try:
        if not os.path.exists(_SCAN_QUEUE_PATH):
            return jsonify({"count": 0, "items": []})
        with open(_SCAN_QUEUE_PATH, encoding='utf-8') as f:
            queue = json.load(f)
        items = [
            {
                "url":    e.get("url", ""),
                "title":  e.get("title", "") or e.get("url", ""),
                "portal": e.get("portal", ""),
                "unanswered": e.get("unanswered", e.get("unanswered_questions", [])),
            }
            for e in queue if e.get("url")
        ]
        return jsonify({"count": len(items), "items": items})
    except Exception as e:
        return jsonify({"count": 0, "items": [], "error": str(e)})


@app.route('/api/scan-queue/dismiss', methods=['POST'])
def api_scan_queue_dismiss():
    """Elimina una URL de la cola de scan."""
    url = (request.json or {}).get('url', '').strip()
    if not url:
        return jsonify({'ok': False}), 400
    try:
        if os.path.exists(_SCAN_QUEUE_PATH):
            with open(_SCAN_QUEUE_PATH, encoding='utf-8') as f:
                queue = json.load(f)
            queue = [e for e in queue if e.get('url') != url]
            with open(_SCAN_QUEUE_PATH, 'w', encoding='utf-8') as f:
                json.dump(queue, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pending-questions')
def api_pending_questions():
    """Devuelve la lista de preguntas pendientes de respuesta del usuario.
    Incluye 'suggested_answer' si hay coincidencia en qa_cache."""
    try:
        if not os.path.exists(_PENDING_Q_PATH):
            return jsonify([])
        with open(_PENDING_Q_PATH, encoding='utf-8') as f:
            data = json.load(f)
        # Enriquecer con sugerencia del cache para preguntas sin respuesta
        for entry in data:
            if not entry.get('answered') and not entry.get('answer'):
                norm = entry.get('norm', '')
                if norm:
                    entry['suggested_answer'] = _suggest_from_cache(norm)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pending-questions/auto-fill', methods=['POST'])
def api_auto_fill_questions():
    """
    Auto-rellena preguntas pendientes usando (en orden de prioridad):
      1. question_answers.json + qa_cache.json
      2. profile_kb.json qa_overrides
      3. Variables USER_* del .env / config

    Retorna detalle de cuántas se llenaron y desde qué fuente.
    """
    try:
        if not os.path.exists(_PENDING_Q_PATH):
            return jsonify({"filled": 0, "remaining": 0, "details": []})
        with open(_PENDING_Q_PATH, encoding='utf-8') as f:
            pending = json.load(f)

        # Cargar QA principal para persistir nuevas respuestas
        qa: dict = {}
        if os.path.exists(_QA_PATH):
            try:
                with open(_QA_PATH, encoding='utf-8') as f:
                    qa = json.load(f)
            except Exception:
                pass

        filled = 0
        details = []
        for entry in pending:
            if entry.get('answered'):
                continue
            norm = entry.get('norm', '')
            if not norm:
                continue
            suggestion = _suggest_from_cache(norm)
            if suggestion:
                entry['answer']   = suggestion
                entry['answered'] = True
                entry['auto_answered'] = True   # marcar como auto-respondido
                filled += 1
                qa[norm] = suggestion           # aprender para futuros runs
                details.append({
                    "label": entry.get('label', norm)[:60],
                    "answer": suggestion[:80],
                })

        os.makedirs(os.path.dirname(_PENDING_Q_PATH), exist_ok=True)
        with open(_PENDING_Q_PATH, 'w', encoding='utf-8') as f:
            json.dump(pending, f, ensure_ascii=False, indent=2)
        with open(_QA_PATH, 'w', encoding='utf-8') as f:
            json.dump(qa, f, ensure_ascii=False, indent=2)

        remaining = sum(1 for e in pending if not e.get('answered'))
        return jsonify({"ok": True, "filled": filled, "remaining": remaining, "details": details})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/pending-questions/answer', methods=['POST'])
def api_answer_question():
    """Guarda la respuesta del usuario para una pregunta pendiente."""
    try:
        payload = request.get_json(force=True)
        norm    = (payload.get('norm') or '').strip().lower()[:200]
        answer  = (payload.get('answer') or '').strip()
        if not norm or not answer:
            return jsonify({"ok": False, "error": "norm y answer son requeridos"}), 400

        existing = []
        if os.path.exists(_PENDING_Q_PATH):
            with open(_PENDING_Q_PATH, encoding='utf-8') as f:
                existing = json.load(f)

        updated = False
        for entry in existing:
            if entry.get('norm', '') == norm:
                entry['answer']   = answer
                entry['answered'] = True
                updated = True
                break

        if not updated:
            return jsonify({"ok": False, "error": "Pregunta no encontrada"}), 404

        os.makedirs(os.path.dirname(_PENDING_Q_PATH), exist_ok=True)
        with open(_PENDING_Q_PATH, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)

        # Persistir también en question_answers.json para auto-fill en futuros runs
        qa: dict = {}
        if os.path.exists(_QA_PATH):
            try:
                with open(_QA_PATH, encoding='utf-8') as f:
                    qa = json.load(f)
            except Exception:
                pass
        qa[norm] = answer
        with open(_QA_PATH, 'w', encoding='utf-8') as f:
            json.dump(qa, f, ensure_ascii=False, indent=2)

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/save-qa', methods=['POST'])
def api_save_qa():
    """Guarda directamente una respuesta en question_answers.json (sin necesitar pending list)."""
    try:
        payload = request.get_json(force=True)
        norm    = (payload.get('norm') or '').strip().lower()[:200]
        answer  = (payload.get('answer') or '').strip()
        if not norm or not answer:
            return jsonify({"ok": False, "error": "norm y answer requeridos"}), 400

        qa: dict = {}
        if os.path.exists(_QA_PATH):
            try:
                with open(_QA_PATH, encoding='utf-8') as f:
                    qa = json.load(f)
            except Exception:
                pass
        qa[norm] = answer
        os.makedirs(os.path.dirname(_QA_PATH), exist_ok=True)
        with open(_QA_PATH, 'w', encoding='utf-8') as f:
            json.dump(qa, f, ensure_ascii=False, indent=2)

        # También marcar como respondida en pending si existe
        if os.path.exists(_PENDING_Q_PATH):
            try:
                with open(_PENDING_Q_PATH, encoding='utf-8') as f:
                    pending = json.load(f)
                for entry in pending:
                    if entry.get('norm') == norm:
                        entry['answered'] = True
                        entry['answer']   = answer
                with open(_PENDING_Q_PATH, 'w', encoding='utf-8') as f:
                    json.dump(pending, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@socketio.on('connect', namespace='/bot')
def handle_connect():
    print("Cliente conectado al canal /bot")
    emit('bot_status', state.get_status())
    emit('session_status', get_session_status())

@socketio.on('start_master', namespace='/bot')
def handle_start(data):
    """Botón Postular → postulación directa independiente (no necesita scan previo)."""
    _log.info("start_master recibido: portals=%s apply_active=%s", data.get('portals', []), state.apply_active)
    if state.apply_active:
        # Verificar si el proceso aún vive — si murió, limpiar el estado colgado
        with state.lock:
            proc = state.apply_process
            if proc is None or proc.poll() is not None:
                _log.warning("apply_active=True pero proceso muerto — limpiando estado.")
                state.apply_active  = False
                state.apply_process = None
            else:
                state.add_log("\n[SISTEMA] ⚠️ Ya hay una postulación activa. Detén el bot antes de iniciar otra.\n")
                emit('bot_status', state.get_status() | {"status": "already_running"})
                return
    portals = _validate_portals(data.get('portals', []))
    if not portals:
        # Sin portales seleccionados → usar todos los habilitados por defecto
        portals = [p for p in _KNOWN_PORTALS if p != 'indeed']
        state.add_log("\n[SISTEMA] Sin portales seleccionados — usando todos los disponibles.\n")
    runtime_env = {
        key: clean_form_value(value)
        for key, value in (data.get('profile') or {}).items()
        if key in _SENSITIVE_ENV_KEYS and clean_form_value(value)
    }
    if data.get('keywords'):
        runtime_env['USER_KEYWORDS'] = clean_form_value(data.get('keywords'))
    if data.get('max_offers'):
        runtime_env['USER_MAX_OFFERS'] = clean_form_value(data.get('max_offers'))

    persistent = bool(data.get('persistent', True))  # persistente por defecto
    try:
        min_per = max(1, min(50, int(str(data.get('min_per_portal', 1)).strip())))
    except (ValueError, TypeError):
        min_per = 1

    def _run():
        run_id = state.start_apply()
        socketio.emit('bot_status', state.get_status() | {"status": "started"}, namespace='/bot')

        def _final_finish():
            # Resumen de sesión con conteo final
            from bot.state import get_stats as _get_stats
            _s = _get_stats()
            _by_p = _s.get("by_portal", {})
            _applied_total = sum(v.get("applied", 0) for v in _by_p.values())
            _ext_total     = sum(
                sum(cnt for k, cnt in v.items() if k.startswith("external"))
                for v in _by_p.values()
            )
            _err_total = sum(
                sum(cnt for k, cnt in v.items() if k.startswith("error"))
                for v in _by_p.values()
            )
            state.add_log(
                f"\n[POSTULAR] ✅ Sesión completada — "
                f"{_applied_total} postuladas · {_ext_total} externas · {_err_total} errores\n"
            )
            # Enviar notificación por email/webhook si está configurado
            try:
                from bot.notifier import send_summary as _send_summary
                _run_start_ts = time.time()  # aproximado — no tenemos start exacto aquí
                _send_summary(
                    portals=portals,
                    applied=_applied_total,
                    external=_ext_total,
                    filtered=0,
                    errors=_err_total,
                )
            except Exception as _ne:
                _log.debug("[NOTIFIER] Error enviando resumen: %s", _ne)
            if state.finish_apply(run_id):
                socketio.emit('bot_status', state.get_status() | {"status": "finished"}, namespace='/bot')
                socketio.emit('session_status', get_session_status(), namespace='/bot')

        try:
            # ── Paso 1: Búsqueda multi-keyword / persistente ─────────────────
            if persistent:
                cmd = [sys.executable, "-u", "main.py",
                       "--persistent", "--headless",
                       "--portal", ",".join(portals),
                       "--min-per-portal", str(min_per)]
            else:
                cmd = [sys.executable, "-u", "main.py", "--multi-keyword", "--headless",
                       "--portal", ",".join(portals)]
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, encoding='utf-8', errors='replace', bufsize=1,
                                    env=_make_child_env(runtime_env))
            state.set_apply_process(proc)
            _start_watchdog(proc, MASTER_TIMEOUT_S, f"master-{','.join(portals)}")
            # _stream_process verifica stop_requested cada 200ms — garantiza stop responsivo
            _stream_process(proc, "\n[POSTULAR] Búsqueda con keywords completada.\n", lambda: None)

            # ── Paso 2: Procesar cola de scan acumulado ───────────────────────
            if not state.stop_requested:
                try:
                    if os.path.exists(_SCAN_QUEUE_PATH):
                        with open(_SCAN_QUEUE_PATH, encoding='utf-8') as _qf:
                            _queue = json.load(_qf)
                        if _queue:
                            state.add_log(
                                f"\n[APPLY-QUEUE] ⏳ {len(_queue)} ofertas en cola "
                                f"— procesando ahora...\n"
                            )
                            cmd2 = [sys.executable, "-u", "main.py", "--apply-queue", "--headless",
                                    "--portal", ",".join(portals)]
                            proc2 = subprocess.Popen(
                                cmd2, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                text=True, encoding='utf-8', errors='replace', bufsize=1,
                                env=_make_child_env(runtime_env)
                            )
                            state.set_apply_process(proc2)
                            _start_watchdog(proc2, PORTAL_TIMEOUT_S,
                                            f"apply-queue-{','.join(portals)}")
                            _stream_process(proc2, "\n[APPLY-QUEUE] Cola procesada.\n",
                                            lambda: None)
                except Exception as _qe:
                    state.add_log(f"\n[APPLY-QUEUE] Error procesando cola: {_qe}\n")
        except Exception as e:
            state.add_log(f"\n[POSTULAR] Error crítico: {e}\n")
        finally:
            _final_finish()

    threading.Thread(target=_run, daemon=True).start()
    socketio.emit('bot_status', state.get_status() | {"status": "starting"}, namespace='/bot')

def _do_stop():
    """Lógica compartida de stop (usada por HTTP y SocketIO)."""
    killed = state.stop_process()
    state.add_log("\n[SISTEMA] Deteniendo ejecución a solicitud del usuario.\n")
    def _emit_stopped():
        time.sleep(0.8)
        # Limpiar stop_requested para que el próximo run no quede pegado
        with state.lock:
            state.stop_requested = False
        socketio.emit('bot_status', state.get_status() | {"status": "stopped"}, namespace='/bot')
    threading.Thread(target=_emit_stopped, daemon=True).start()
    return killed


@socketio.on('stop_master', namespace='/bot')
def handle_stop():
    _do_stop()


@app.route('/api/reset-all', methods=['POST'])
def api_reset_all():
    """
    Limpia todo el estado persistente:
    - Cola de scan (scan_queue.json)
    - Quick links (quick_links.json)
    - Restricciones de portales (portal_restrictions.json)
    - Stats del optimizer de keywords (keyword_stats.json)
    - Señal de stop (STOP_SIGNAL) si quedó colgada
    No toca: .env, sessions/, uploads/, logs/, ni la DB de postulaciones.
    """
    if state.is_active:
        return jsonify({'ok': False, 'msg': 'Detén el bot antes de limpiar.'}), 409

    cleared = []
    errors  = []

    _files = [
        (_SCAN_QUEUE_PATH,   'Cola de scan'),
        (_QUICK_LINKS_PATH,  'Quick links'),
        (_RESTRICTIONS_PATH, 'Restricciones de portales'),
        (STOP_SIGNAL_PATH,   'Señal de stop'),
        (os.path.join(_DATA_DIR, 'keyword_stats.json'), 'Stats de keywords'),
    ]
    for path, label in _files:
        try:
            if os.path.exists(path):
                os.remove(path)
                cleared.append(label)
        except Exception as e:
            errors.append(f'{label}: {e}')

    state.clear_logs()
    socketio.emit('bot_status', state.get_status() | {'status': 'reset'}, namespace='/bot')
    return jsonify({'ok': True, 'cleared': cleared, 'errors': errors})


@app.route('/api/daily-stats', methods=['GET'])
def api_daily_stats():
    """
    Retorna conteos agrupados por fecha para los últimos N días.
    Fuente 1: archivos logs/applied_YYYY-MM-DD.csv (persistente entre sesiones)
    Fuente 2: bot.state en memoria (sesión actual, cubre gaps no escritos aún a disco)
    Response: { ok, labels, applied, external, skipped }
    """
    import csv as _csv
    from datetime import date as _date, timedelta
    from pathlib import Path as _Path
    from collections import defaultdict

    days = int(request.args.get('days', 14))
    days = min(max(days, 1), 90)

    end   = _date.today()
    start = end - timedelta(days=days - 1)

    by_day: dict = defaultdict(lambda: {'applied': 0, 'external': 0, 'skipped': 0})

    # ── Fuente 1: CSV logs en disco ──────────────────────────────────────────
    logs_dir = _Path(__file__).parent / 'logs'
    cur = start
    while cur <= end:
        csv_path = logs_dir / f"applied_{cur.strftime('%Y-%m-%d')}.csv"
        if csv_path.exists():
            try:
                with open(csv_path, newline='', encoding='utf-8', errors='replace') as f:
                    reader = _csv.DictReader(f)
                    for row in reader:
                        st  = (row.get('status') or '').strip()
                        day = (row.get('timestamp') or '')[:10]
                        if not day:
                            continue
                        if st == 'applied':
                            by_day[day]['applied'] += 1
                        elif st.startswith('external'):
                            by_day[day]['external'] += 1
                        elif st.startswith('skipped') or st == 'dry_run' or st == 'filtered':
                            by_day[day]['skipped'] += 1
            except Exception:
                pass
        cur += timedelta(days=1)

    # ── Fuente 2: memoria (sesión actual — evita duplicar si ya se escribió a disco) ──
    # Solo suma registros cuya fecha NO aparece en disco (no hay CSV para ese día aún)
    dates_from_disk = set(by_day.keys())
    try:
        from bot.state import get_recent
        for r in get_recent(limit=10000):
            day = (r.get('applied_at') or '')[:10]
            if not day or day in dates_from_disk:
                continue
            st = r.get('status', '')
            if st == 'applied':
                by_day[day]['applied'] += 1
            elif st.startswith('external'):
                by_day[day]['external'] += 1
            elif st.startswith('skipped') or st == 'dry_run':
                by_day[day]['skipped'] += 1
    except Exception:
        pass

    # ── Construir arrays para el rango ──────────────────────────────────────
    labels, applied_vals, external_vals, skipped_vals = [], [], [], []
    cur = start
    while cur <= end:
        ds  = cur.strftime('%Y-%m-%d')
        row = by_day.get(ds, {})
        labels.append(ds)
        applied_vals.append(row.get('applied', 0))
        external_vals.append(row.get('external', 0))
        skipped_vals.append(row.get('skipped', 0))
        cur += timedelta(days=1)

    return jsonify({
        'ok':      True,
        'labels':  labels,
        'applied':  applied_vals,
        'external': external_vals,
        'skipped':  skipped_vals,
    })


@app.route('/api/export-csv', methods=['GET'])
def api_export_csv():
    """Descarga las postulaciones de la sesión actual como Excel (.xlsx)."""
    import io
    from datetime import date as _date
    from flask import Response as _Resp
    from bot.state import get_recent
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        rows = get_recent(limit=10000)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postulaciones"

    # Estilos
    header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill   = PatternFill('solid', fgColor='1A3A5C')
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    even_fill = PatternFill('solid', fgColor='EFF6FF')

    # Status → color de celda
    status_colors = {
        'applied':   '00C853',   # verde
        'external':  '0288D1',   # azul
        'filtered':  'FFB300',   # amarillo
        'error':     'E53935',   # rojo
        'no_nav':    'F4511E',   # naranja
    }

    # Encabezados
    headers = ['#', 'Portal', 'Título', 'Estado', 'Fecha / Hora', 'URL']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # Filas de datos
    for i, r in enumerate(rows, 1):
        status = r.get('status', '')
        ws.append([
            i,
            r.get('portal', ''),
            r.get('title', ''),
            status,
            r.get('applied_at', ''),
            r.get('url', ''),
        ])
        row_idx = i + 1
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = thin_border
            cell.alignment = center_align if col_idx != 3 else left_align
            if i % 2 == 0:
                cell.fill = even_fill

        # Color de estado
        status_cell = ws.cell(row=row_idx, column=4)
        color = status_colors.get(status)
        if color:
            status_cell.fill = PatternFill('solid', fgColor=color)
            status_cell.font = Font(color='FFFFFF', bold=True, size=10)

        # URL como hipervínculo
        url = r.get('url', '')
        if url:
            url_cell = ws.cell(row=row_idx, column=6)
            url_cell.hyperlink = url
            url_cell.font = Font(color='0563C1', underline='single')

    # Anchos de columna
    col_widths = [5, 14, 48, 12, 20, 50]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Altura de filas
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 18

    # Freeze header + filtros automáticos
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"postulaciones_{_date.today().strftime('%Y-%m-%d')}.xlsx"
    return _Resp(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/stop', methods=['POST'])
def api_stop():
    """Endpoint HTTP de stop — fallback por si el SocketIO no llega."""
    killed = _do_stop()
    return jsonify({'ok': True, 'killed': killed})


@app.route('/api/reset-sessions', methods=['POST'])
def api_reset_sessions():
    """Borra cookies de todos los portales para forzar re-login en el próximo run."""
    _clear_session_auth()
    return jsonify({'ok': True, 'msg': 'Sesiones borradas — el bot pedirá login en cada portal'})


@app.route('/api/login-portals', methods=['POST'])
def api_login_portals():
    """Abre browser visible para login manual en los portales indicados."""
    data    = request.get_json(force=True) or {}
    portals = data.get('portals', [])
    if not portals:
        return jsonify({'ok': False, 'error': 'Sin portales'}), 400

    # Bloquear SCAN/POSTULAR antes del thread — evita race condition
    with state.lock:
        state.apply_active = True

    def _run_login():
        # Detener bot activo dentro del thread — no bloquea el HTTP handler
        if state.scan_process and state.scan_process.poll() is None:
            state.stop_process()
            time.sleep(1.5)

        try:
            cmd = [sys.executable, "-u", "login_portals.py"] + portals
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True, encoding='utf-8', errors='replace',
                bufsize=1,
                env=_make_child_env(),
            )
            with state.lock:
                state.apply_process = proc
            socketio.emit('bot_status', {'status': 'login_started', 'portals': portals}, namespace='/bot')
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    state.add_log(line)
            proc.wait()
        except Exception as e:
            state.add_log(f"[LOGIN] Error: {e}")
        finally:
            with state.lock:
                state.apply_active  = False
                state.apply_process = None
            socketio.emit('bot_status', {'status': 'login_finished'}, namespace='/bot')
            socketio.emit('session_status', get_session_status(), namespace='/bot')

    t = threading.Thread(target=_run_login, daemon=True)
    t.start()
    return jsonify({'ok': True})

def _on_exit():
    """Limpieza al cerrar el servidor (Ctrl+C o kill)."""
    _log.info("[SHUTDOWN] Servidor cerrándose — limpiando procesos...")
    _write_stop_signal()
    for proc_attr in ('apply_process', 'scan_process'):
        proc = getattr(state, proc_attr, None)
        if proc and proc.poll() is None:
            try:
                proc.terminate()
                proc.wait(timeout=5)
            except Exception:
                try:
                    proc.kill()
                except Exception:
                    pass
    _clear_stop_signal()
    # Borrar cookies de sesión al cerrar — las sesiones solo duran mientras el servidor corre.
    # Al reiniciar el servidor habrá que volver a loguear.
    try:
        _clear_session_auth()
        _log.info("[SHUTDOWN] Sesiones de portales borradas.")
    except Exception as _e:
        _log.debug("[SHUTDOWN] Error borrando sesiones: %s", _e)
    _log.info("[SHUTDOWN] Limpieza completada.")

atexit.register(_on_exit)


# ── Scheduler ─────────────────────────────────────────────────────────────────
# Lee SCHEDULE_PORTALS y SCHEDULE_TIMES del .env y dispara el bot
# automáticamente a las horas configuradas si no hay una ejecución activa.
#
# SCHEDULE_PORTALS=linkedin,computrabajo
# SCHEDULE_TIMES=09:00,14:00,18:00
#
# El scheduler corre en un thread daemon — se inicia al arrancar el servidor y
# muere cuando el proceso principal termina. No persiste entre reinicios.
# Si el bot ya está corriendo a la hora programada, la dispara se omite silenciosamente.

def _scheduler_thread():
    """Thread daemon que dispara el bot a las horas configuradas.
    Re-lee SCHEDULE_PORTALS y SCHEDULE_TIMES en cada ciclo para reflejar
    cambios en .env sin necesidad de reiniciar el servidor.
    """
    import datetime as _dt

    _last_fired: set[tuple] = set()   # (date, HH, MM) disparados hoy

    while True:
        try:
            # Re-leer configuración en cada ciclo (soporta cambios en .env en caliente)
            load_dotenv(override=True)
            _SCHEDULE_PORTALS_RAW = os.getenv("SCHEDULE_PORTALS", "").strip()
            _SCHEDULE_TIMES_RAW   = os.getenv("SCHEDULE_TIMES",   "").strip()

            if not _SCHEDULE_PORTALS_RAW or not _SCHEDULE_TIMES_RAW:
                time.sleep(30)
                continue

            # Parsear portales
            _VALID_SCHED = set(_KNOWN_PORTALS) | {'indeed'}
            sched_portals = [p.strip().lower() for p in _SCHEDULE_PORTALS_RAW.split(",")
                             if p.strip().lower() in _VALID_SCHED]
            if not sched_portals:
                time.sleep(30)
                continue

            # Parsear horarios HH:MM
            sched_times: list[tuple[int, int]] = []
            for t in _SCHEDULE_TIMES_RAW.split(","):
                t = t.strip()
                try:
                    h, m = t.split(":")
                    sched_times.append((int(h), int(m)))
                except (ValueError, AttributeError):
                    pass

            if not sched_times:
                time.sleep(30)
                continue

            now = _dt.datetime.now()
            today = now.date()

            for h, m in sched_times:
                key = (today, h, m)
                if key in _last_fired:
                    continue   # ya se disparó hoy a esta hora

                # Ventana de ±30 s (mitad del intervalo de sleep) — evita doble disparo
                target = now.replace(hour=h, minute=m, second=0, microsecond=0)
                diff   = abs((now - target).total_seconds())
                if diff <= 30:
                    if state.is_active:
                        _log.info("[SCHEDULER] %02d:%02d — bot ya activo, omitiendo.",
                                  h, m)
                        _last_fired.add(key)  # marcar igual para no reintentar este minuto
                    else:
                        _log.info("[SCHEDULER] Disparando bot programado %02d:%02d — %s",
                                  h, m, ", ".join(sched_portals))
                        print(f"\n[SCHEDULER] ⏰ {h:02d}:{m:02d} — lanzando bot: "
                              f"{', '.join(p.upper() for p in sched_portals)}")
                        state.add_log(
                            f"\n[SCHEDULER] ⏰ {h:02d}:{m:02d} — ejecución programada iniciada.\n"
                        )
                        threading.Thread(
                            target=run_bot_thread,
                            args=(sched_portals,),
                            daemon=True,
                        ).start()
                        _last_fired.add(key)  # marcar solo tras lanzar el thread

            # Limpiar disparos de días anteriores (evitar que el set crezca)
            _last_fired = {k for k in _last_fired if k[0] == today}

        except Exception as _se:
            _log.warning("[SCHEDULER] Error en ciclo: %s", _se)

        time.sleep(30)   # comprobar cada 30 segundos


# Iniciar scheduler al arrancar el módulo (funciona tanto en __main__ como importado)
threading.Thread(target=_scheduler_thread, daemon=True, name="scheduler").start()


def _clear_session_auth() -> None:
    """
    Borra datos de login de cada portal al iniciar el servidor.
    Playwright guarda el perfil en user_data_dir con esta estructura real:
      <portal>/                        ← raíz (algunos archivos aquí)
        Login Data, Web Data, LOCK…
        Default/
          Login Data, Web Data…
          Network/
            Cookies, Cookies-journal   ← cookies reales
          Local Storage/
          Session Storage/
          IndexedDB/
          Sessions/
    Solo se eliminan archivos de auth; se preservan model files y caché de Chromium.
    """
    import shutil as _shutil
    from pathlib import Path as _Path

    # Rutas relativas a cada portal_dir que contienen auth
    # Formato: ("archivo_o_dir", es_dir)
    AUTH_TARGETS_ROOT = [
        ("Login Data", False), ("Login Data-journal", False),
        ("Login Data For Account", False), ("Login Data For Account-journal", False),
        ("Web Data", False), ("Web Data-journal", False),
        ("Local Storage", True), ("Session Storage", True),
        ("IndexedDB", True), ("LOCK", False),
    ]
    AUTH_TARGETS_DEFAULT = [
        ("Login Data", False), ("Login Data-journal", False),
        ("Login Data For Account", False), ("Login Data For Account-journal", False),
        ("Web Data", False), ("Web Data-journal", False),
        ("Local Storage", True), ("Session Storage", True),
        ("IndexedDB", True), ("Sessions", True),
        ("SharedStorage", False),
    ]
    AUTH_TARGETS_DEFAULT_NETWORK = [
        ("Cookies", False), ("Cookies-journal", False),
        ("Trust Tokens", False), ("Trust Tokens-journal", False),
    ]

    sessions_dir = _Path(__file__).parent / "sessions"
    if not sessions_dir.exists():
        return

    portals_cleared = set()

    def _rm(path: _Path, is_dir: bool) -> bool:
        if not path.exists():
            return False
        try:
            if is_dir:
                _shutil.rmtree(path, ignore_errors=True)
            else:
                path.unlink(missing_ok=True)
            return True
        except Exception:
            return False

    for portal_dir in sessions_dir.iterdir():
        if not portal_dir.is_dir():
            continue
        hit = False
        # Raíz del perfil
        for name, is_dir in AUTH_TARGETS_ROOT:
            if _rm(portal_dir / name, is_dir):
                hit = True
        # Default/
        default = portal_dir / "Default"
        for name, is_dir in AUTH_TARGETS_DEFAULT:
            if _rm(default / name, is_dir):
                hit = True
        # Default/Network/
        net = default / "Network"
        for name, is_dir in AUTH_TARGETS_DEFAULT_NETWORK:
            if _rm(net / name, is_dir):
                hit = True
        if hit:
            portals_cleared.add(portal_dir.name)

    if portals_cleared:
        print(f"[SESSION] Login limpiado en {len(portals_cleared)} portal(es): "
              + ", ".join(sorted(portals_cleared)))
    else:
        print("[SESSION] No había datos de login previos.")


if __name__ == '__main__':
    port = 5000
    # _clear_session_auth()  # deshabilitado — sesiones persisten entre reinicios
    print(f"\n--- Iniciando modo maestro (producción) ---")
    print(f"URL: http://127.0.0.1:{port}")
    print(f"Servidor: SocketIO (Auto-detect)")
    print(f"-------------------------------------------\n")

    app.jinja_env.auto_reload = True
    socketio.run(app, host='127.0.0.1', port=port, debug=False, use_reloader=False, allow_unsafe_werkzeug=True)
