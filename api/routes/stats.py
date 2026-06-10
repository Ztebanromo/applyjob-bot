"""api/routes/stats.py — Estadísticas globales, históricas y exportación."""
from __future__ import annotations

from flask import Blueprint, jsonify, request

bp = Blueprint('stats', __name__)


@bp.route('/api/stats')
def api_stats():
    """
    Devuelve estadísticas globales y por portal desde la DB SQLite.
    El dashboard las carga al iniciar para mostrar el historial acumulado.
    """
    try:
        from bot.state import get_stats
        stats = get_stats()
        by_portal = stats.get("by_portal", {})

        # Agregar by_status desde by_portal (get_stats no lo devuelve directamente)
        by_status: dict = {}
        for portal_data in by_portal.values():
            for status, cnt in portal_data.items():
                by_status[status] = by_status.get(status, 0) + cnt

        applied = by_status.get("applied", 0)
        errors  = sum(v for k, v in by_status.items() if "error" in k.lower())
        skipped = sum(v for k, v in by_status.items()
                      if k.startswith("skipped") or k.startswith("external") or k == "dry_run")

        # ── Histórico total (todas las sesiones, desde logs/applied_*.csv) ──
        from bot.verification import get_all_time_total
        all_time = get_all_time_total()

        return jsonify({
            "total":     stats.get("total", 0),
            "applied":   applied,
            "errors":    errors,
            "skipped":   skipped,
            "by_portal": by_portal,
            "by_status": by_status,
            "all_time_applied":   all_time.get("total", 0),
            "all_time_by_portal": all_time.get("by_portal", {}),
        })
    except Exception as e:
        return jsonify({"total": 0, "applied": 0, "errors": 0, "skipped": 0,
                        "by_portal": {}, "by_status": {}, "all_time_applied": 0,
                        "all_time_by_portal": {}, "error": str(e)})


@bp.route('/api/daily-stats', methods=['GET'])
def api_daily_stats():
    """
    Retorna conteos agrupados por fecha para los últimos N días.
    Fuente 1: archivos logs/applied_YYYY-MM-DD.csv (persistente entre sesiones)
    Fuente 2: bot.state en memoria (sesión actual, cubre gaps no escritos aún a disco)
    Response: { ok, labels, applied, external, skipped }
    """
    import csv as _csv
    import os
    from datetime import date as _date, timedelta
    from pathlib import Path as _Path
    from collections import defaultdict

    days = int(request.args.get('days', 14))
    days = min(max(days, 1), 90)

    end   = _date.today()
    start = end - timedelta(days=days - 1)

    by_day: dict = defaultdict(lambda: {'applied': 0, 'external': 0, 'skipped': 0})

    # ── Fuente 1: CSV logs en disco ──────────────────────────────────────────
    logs_dir = _Path(__file__).resolve().parent.parent.parent / 'logs'
    cur = start
    while cur <= end:
        csv_path = logs_dir / f"applied_{cur.strftime('%Y-%m-%d')}.csv"
        if csv_path.exists():
            try:
                with open(csv_path, newline='', encoding='utf-8', errors='replace') as f:
                    reader = _csv.DictReader(f)
                    for row in reader:
                        st  = (row.get('status') or '').strip()
                        day = (row.get('timestamp') or '')[:10]
                        if not day:
                            continue
                        if st == 'applied':
                            by_day[day]['applied'] += 1
                        elif st.startswith('external'):
                            by_day[day]['external'] += 1
                        elif st.startswith('skipped') or st == 'dry_run' or st == 'filtered':
                            by_day[day]['skipped'] += 1
            except Exception:
                pass
        cur += timedelta(days=1)

    # ── Fuente 2: memoria (sesión actual — evita duplicar si ya se escribió a disco) ──
    # Solo suma registros cuya fecha NO aparece en disco (no hay CSV para ese día aún)
    dates_from_disk = set(by_day.keys())
    try:
        from bot.state import get_recent
        for r in get_recent(limit=10000):
            day = (r.get('applied_at') or '')[:10]
            if not day or day in dates_from_disk:
                continue
            st = r.get('status', '')
            if st == 'applied':
                by_day[day]['applied'] += 1
            elif st.startswith('external'):
                by_day[day]['external'] += 1
            elif st.startswith('skipped') or st == 'dry_run':
                by_day[day]['skipped'] += 1
    except Exception:
        pass

    # ── Construir arrays para el rango ──────────────────────────────────────
    labels, applied_vals, external_vals, skipped_vals = [], [], [], []
    cur = start
    while cur <= end:
        ds  = cur.strftime('%Y-%m-%d')
        row = by_day.get(ds, {})
        labels.append(ds)
        applied_vals.append(row.get('applied', 0))
        external_vals.append(row.get('external', 0))
        skipped_vals.append(row.get('skipped', 0))
        cur += timedelta(days=1)

    return jsonify({
        'ok':      True,
        'labels':  labels,
        'applied':  applied_vals,
        'external': external_vals,
        'skipped':  skipped_vals,
    })


@bp.route('/api/export-csv', methods=['GET'])
def api_export_csv():
    """Descarga las postulaciones de la sesión actual como Excel (.xlsx)."""
    import io
    from datetime import date as _date
    from flask import Response as _Resp
    from bot.state import get_recent
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        rows = get_recent(limit=10000)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postulaciones"

    # Estilos
    header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill   = PatternFill('solid', fgColor='1A3A5C')
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    even_fill = PatternFill('solid', fgColor='EFF6FF')

    # Status → color de celda
    status_colors = {
        'applied':   '00C853',   # verde
        'external':  '0288D1',   # azul
        'filtered':  'FFB300',   # amarillo
        'error':     'E53935',   # rojo
        'no_nav':    'F4511E',   # naranja
    }

    # Encabezados
    headers = ['#', 'Portal', 'Título', 'Estado', 'Fecha / Hora', 'URL']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # Filas de datos
    for i, r in enumerate(rows, 1):
        status = r.get('status', '')
        ws.append([
            i,
            r.get('portal', ''),
            r.get('title', ''),
            status,
            r.get('applied_at', ''),
            r.get('url', ''),
        ])
        row_idx = i + 1
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = thin_border
            cell.alignment = center_align if col_idx != 3 else left_align
            if i % 2 == 0:
                cell.fill = even_fill

        # Color de estado
        status_cell = ws.cell(row=row_idx, column=4)
        color = status_colors.get(status)
        if color:
            status_cell.fill = PatternFill('solid', fgColor=color)
            status_cell.font = Font(color='FFFFFF', bold=True, size=10)

        # URL como hipervínculo
        url = r.get('url', '')
        if url:
            url_cell = ws.cell(row=row_idx, column=6)
            url_cell.hyperlink = url
            url_cell.font = Font(color='0563C1', underline='single')

    # Anchos de columna
    col_widths = [5, 14, 48, 12, 20, 50]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Altura de filas
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 18

    # Freeze header + filtros automáticos
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"postulaciones_{_date.today().strftime('%Y-%m-%d')}.xlsx"
    return _Resp(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
